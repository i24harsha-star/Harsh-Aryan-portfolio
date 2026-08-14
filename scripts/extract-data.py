#!/usr/bin/env python3
"""Extract the numbers behind Harsh's analyses into JSON the site can render.

The source workbooks are screenshots-of-Excel in their raw form. Rather than
publish spreadsheet screenshots, the site redraws the same numbers in its own
visual language — so this pulls the real values out once, at build-prep time.

Nothing here invents or adjusts a number: values are copied as computed in the
workbooks, rounded only for display.
"""

import json
import os

import numpy as np
import openpyxl

SRC = "/Users/harsharyan/Files/Work/Projects/Portfolio/Website Assets"
OUT = os.path.join(os.path.dirname(__file__), "..", "src", "data")


def income_quality():
    path = os.path.join(
        SRC,
        "Income Quality of Large Cap Chemicals & Petrochemicals Companies",
        "Accessing the Income Quality of Large Cap Chemicals & Petrochemicals Company.xlsx",
    )
    wb = openpyxl.load_workbook(path, data_only=True)
    groups = []

    for sheet in ("Top 5 Companies", "Bottom 5 Companies"):
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        companies = []
        i = 0
        while i < len(rows):
            name = rows[i][1] if len(rows[i]) > 1 else None
            # A company block starts with its name, then years / NP / CFO / ratio.
            if isinstance(name, str) and name.strip() and i + 4 < len(rows):
                years, np_row, cfo_row, ratio_row = (rows[i + j] for j in range(1, 5))
                if str(np_row[1]).strip() == "Net Profit":
                    yrs, npv, cfov, ratios = [], [], [], []
                    for c in range(2, 14):
                        y = years[c]
                        if hasattr(y, "year"):
                            yrs.append(y.year)
                            npv.append(_num(np_row[c]))
                            cfov.append(_num(cfo_row[c]))
                            ratios.append(_num(ratio_row[c]))
                    companies.append(
                        {
                            "name": name.strip(),
                            "years": yrs,
                            "netProfit": npv,
                            "cfo": cfov,
                            "ratio": ratios,
                            # Columns: 2..13 are the 12 fiscal years, 15 holds the
                            # label ("Average"/"Median"/"Difference"), 16 the value.
                            "average": _num(years[16]) if len(years) > 16 else None,
                            "median": _num(np_row[16]) if len(np_row) > 16 else None,
                            "difference": _num(cfo_row[16]) if len(cfo_row) > 16 else None,
                        }
                    )
                    i += 5
                    continue
            i += 1
        groups.append({"label": sheet.replace(" Companies", ""), "companies": companies})

    _write("income-quality.json", {"groups": groups})
    for g in groups:
        print(f"   income-quality: {g['label']} -> {len(g['companies'])} companies")


def monte_carlo(bins=44):
    path = os.path.join(SRC, "Monte Carlo simulation", "Monte Carlo Simulation.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Monte-Carlo Simulation"]

    values = [
        r[5]
        for r in ws.iter_rows(min_row=12, values_only=True)
        if len(r) > 5 and isinstance(r[5], (int, float))
    ]
    a = np.array(values, dtype=float)
    counts, edges = np.histogram(a, bins=bins)

    _write(
        "monte-carlo.json",
        {
            "inputs": {
                "initialInvestment": 100000,
                "meanReturn": -0.000254,
                "stdDev": 0.006825,
                "iterations": len(values),
            },
            "results": {
                "average": round(float(a.mean()), 2),
                "minimum": round(float(a.min()), 2),
                "maximum": round(float(a.max()), 2),
                "stdDev": round(float(a.std()), 2),
            },
            "histogram": {
                "counts": counts.tolist(),
                "edges": [round(float(e), 2) for e in edges],
            },
        },
    )
    print(f"   monte-carlo: {len(values)} iterations -> {bins} bins")


def _num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    return None


def _write(name, payload):
    os.makedirs(OUT, exist_ok=True)
    with open(os.path.join(OUT, name), "w") as f:
        json.dump(payload, f, indent=1)


if __name__ == "__main__":
    income_quality()
    monte_carlo()
    print("done.")
